import openpyxl
import os
# open file and read line by line
repo_path = os.path.dirname(os.path.abspath(__file__))

file1 = os.path.join(repo_path, '..', 'data', 'Collated transaction reports.xlsx')
file2 = os.path.join(repo_path, '..', 'data', 'Knowledge_Schools_Trust_-_Account_Transactions (23).xlsx')

wb1 = openpyxl.load_workbook(file1)
ws1 = wb1.active

wb2 = openpyxl.load_workbook(file2)
ws2 = wb2.active

for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):
    for cell in row:
        print(cell.value)

for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=ws2.max_column):
    for cell in row:
        print(cell.value)   