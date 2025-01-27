import pandas as pd
import os
import sys
import openpyxl

from openpyxl.styles import PatternFill

def compare_and_highlight_excel(file1, file2, sheet1=0, sheet2=0):
    # Read Excel files
    df1 = pd.read_excel(file1, sheet_name=sheet1, header=3)
    df2 = pd.read_excel(file2, sheet_name=sheet2, header=6)
    print(df1.head(5))
    # # Find matching columns
    # common_columns = list(set(df1.columns) & set(df2.columns))

    common_columns = ['Date', 'Description', 'Debit']

    # Merge dataframes
    merged = pd.merge(df1, df2, 
                    on=common_columns, 
                    how='outer', 
                    indicator=True)
#   Source  ...  VAT Unnamed: 10  Unnamed: 11
# 0        NaN                                                NaN              NaN  ...  NaN         NaN          NaN
# 1        NaN  Supplies & Services Costs: Educational Supplie...              NaN  ...  NaN         NaN          NaN
# 2        Oct                                2024-10-01 00:00:00  Payable Invoice  ...  0.0         NaN          NaN
# 3        Oct                                2024-10-01 00:00:00  Payable Invoice  ...  0.0         NaN          NaN
# 4        Oct                                2024-10-01 00:00:00  Payable Invoice  ...  0.0         NaN          NaN
    
    # Identify unmatched transactions
    unmatched_table1 = merged[merged['_merge'] == 'left_only']
#    print(unmatched_table1.head(5))
#    Unnamed: 0                 Date             Source_x  ... Gross_y VAT_y     _merge
#203        Oct  2024-10-30 00:00:00                  NaN  ...     NaN   NaN  left_only
#235        Oct  2024-09-19 00:00:00      Payable Invoice  ...     NaN   NaN  left_only
#250        Nov  2024-10-04 00:00:00  Payable Credit Note  ...     NaN   NaN  left_only
#251        Nov  2024-10-04 00:00:00  Payable Credit Note  ...     NaN   NaN  left_only
#255        Nov  2024-10-31 00:00:00                  NaN  ...     NaN   NaN  left_only
    unmatched_table2 = merged[merged['_merge'] == 'right_only']
#    print(unmatched_table2.head(5))
#    Unnamed: 0                                           Date Source_x  ...   Gross_y VAT_y      _merge
#543        NaN        Income: DfE Revenue Grants: English Hub      NaN  ...       NaN   NaN  right_only
#544        NaN                            2024-09-27 00:00:00      NaN  ...   8609.48   0.0  right_only
#545        NaN                            2024-10-31 00:00:00      NaN  ...  47809.03   0.0  right_only
#546        NaN                            2024-11-29 00:00:00      NaN  ...  20049.35   0.0  right_only
#547        NaN  Total Income: DfE Revenue Grants: English Hub      NaN  ...  76467.86   0.0  right_only


    # Load workbooks
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)

    # Highlight unmatched in first file
    ws1 = wb1.worksheets[sheet1]
    ws2 = wb2.worksheets[sheet2]
    yelow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
    # Fantasy fill
    red_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
    # Fantasy Fill
    # for row in ws1.iter_rows(min_row=3):  # Start from header row
    #     if any(all(row[df1.columns.get_loc(col)].value == unmatched_table1.iloc[i][col] for col in common_columns) for i in range(len(unmatched_table1))):
    #         for cell in row:
    #             cell.fill = yelow_fill

    # for row in ws1.iter_rows(min_row=3):  # Start from header row
        # if any ___ ( all ___ ( true_or_false ___ for col in common_columns) ___ for unmatched_row in range(len(unmatched_table1))):
        # we want to slice the above expression into multiple parts to understand it better
        # for i in range(len(unmatched_table1)):
        #     all_cols_match = True
        #     for col in common_columns:
        #         if row[df1.columns.get_loc(col)].value != unmatched_table1.iloc[i][col]:
        #             all_cols_match = False
        #             break
        #     if all_cols_match:
        #         matches = True
        #         break
        # if matches:
        #     for cell in row:
        #         cell.fill = yelow_fill
        # we can see that we are checking if the row is in unmatched_table1
        # if it is we are highlighting the row
        # so we can simplify the above code to
# ___col___ = Date, Description, Debit
# ws1 
# (row[df1.columns.get_loc(___col___)].value == unmatched_table1.iloc[___unmatched_row___][____col____])
       
        # if one_match_found:
        #     for cell in row:
        #        cell.fill = yelow_fill

    # any: Return True if bool(x) is True for any x in the iterable (dicts, lists, sets, etc.)
    # If the iterable is empty, return False.
    # dict is a collection which is unordered, changeable and indexed. In Python dictionaries are written with curly brackets, and they have keys and values.
    # list is a collection which is ordered and changeable. Allows duplicate members.
    # set is a collection which is unordered and unindexed. No duplicate members.
    # e.g ['a', 'b', 'c'] 
    # if we say any(x == 'a' for x in ['a', 'b', 'c']) it will return True
    # if we say all(x == 'a' for x in ['a', 'b', 'c']) it will return False
    # but for example ['a', 'a', 'a']
    # if we say all(x == 'a' for x in ['a', 'a', 'a']) it will return True

# Loop through each row in worksheet, starting at row 3
# for row in ws1.iter_rows(min_row=3):

# explain___________________________
    # Check if current row matches any unmatched transaction by comparing all columns
    # matches = False
    # for i in range(len(unmatched_table1)):
    #     all_cols_match = True
    #     for col in common_columns:
    #         if row[df1.columns.get_loc(col)].value != unmatched_table1.iloc[i][col]:
    #             all_cols_match = False
    #             break
    #     if all_cols_match:
    #         matches = True
    #         break

    # # If row matches an unmatched transaction, highlight all cells yellow
    # if matches:
    #     for cell in row:
    #         cell.fill = yellow_fill
    # end of explain________________________
                        
    # for row in ws2.iter_rows(min_row=6):
    #     if any(all(row[df2.columns.get_loc(col)].value == unmatched_table2.iloc[i][col] for col in common_columns) for i in range(len(unmatched_table2))):
    #         for cell in row:
    #             cell.fill = yelow_fill

    
    # Save highlighted workbook
    # wb1.save('highlighted_table1.xlsx')
    # wb2.save('highlighted_table2.xlsx')
    
    # print(f"Unmatched in Table 1: {len(unmatched_table1)}")
    # print(f"Unmatched in Table 2: {len(unmatched_table2)}")

# Get script directory
repo_path = os.path.dirname(os.path.abspath(__file__))

# Construct file paths
file1 = os.path.join(repo_path, '..', 'data', 'Collated transaction reports.xlsx')
file2 = os.path.join(repo_path, '..', 'data', 'Knowledge_Schools_Trust_-_Account_Transactions (23).xlsx')

# Run comparison and highlighting
compare_and_highlight_excel(file1, file2)