# import pandas as pd
# # import numpy as np
# import os

# import os

# # Absolute path method
# file1 = '/Users/ekaterina/Desktop/coding/reconciliation_automation/data/Collated transaction reports.xlsx'
# file2 = '/Users/ekaterina/Desktop/coding/reconciliation_automation/data/Knowledge_Schools_Trust_-_Account_Transactions (23).xlsx'

# # Relative path method (recommended) - later look into that, this is not horosh to have it as above.
# # repo_path = os.path.dirname(os.path.abspath(__file__))
# # file1 = os.path.join(repo_path, 'financial_data', 'table1.xlsx')
# # file2 = os.path.join(repo_path, 'financial_data', 'table2.xlsx')

# def compare_excel_tables(file1, file2, sheet1=0, sheet2=0):
#     # Read Excel files with headers on line 4 and 7
#     df1 = pd.read_excel(file1, sheet_name=sheet1, header=3)
#     df2 = pd.read_excel(file2, sheet_name=sheet2, header=6)

#     # # Print header if we get the right ones
#     # print('Headers Table 1:', list(df1.columns))
#     # print("header Table 2:", list(df2.columns))
    
#     # Find matching columns
#     common_columns = list(set(df1.columns) & set(df2.columns))
    
#     # Merge dataframes
#     merged = pd.merge(df1, df2, 
#                     on=compare_columns, 
#                     how='outer', 
#                     indicator=True)
    
#     # Add match status column
#     merged['Match_Status'] = merged['_merge'].map({
#         'left_only': 'Only in Table 1',
#         'right_only': 'Only in Table 2',
#         'both': 'Match'
#     })
    
#     # Optional: Save results
#     merged.to_excel('comparison_results.xlsx', index=False)
    
#     return merged

# # Example usage
# result = compare_excel_tables(file1, file2)
# print(result['Match_Status'].value_counts())

import pandas as pd
import os
import openpyxl
from openpyxl.styles import PatternFill

def compare_and_highlight_excel(file1, file2, sheet1=0, sheet2=0):
    # Read Excel files
    df1 = pd.read_excel(file1, sheet_name=sheet1, header=3)
    df2 = pd.read_excel(file2, sheet_name=sheet2, header=6)
    
    # Find matching columns
    common_columns = list(set(df1.columns) & set(df2.columns))
    
    # Merge dataframes
    merged = pd.merge(df1, df2, 
                    on=common_columns, 
                    how='outer', 
                    indicator=True)
    
    # Identify unmatched transactions
    unmatched_table1 = merged[merged['_merge'] == 'left_only']
    unmatched_table2 = merged[merged['_merge'] == 'right_only']
    
    # Load workbooks
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)
    
    # Highlight unmatched in first file
    ws1 = wb1.worksheets[sheet1]
    yelow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
    
    for row in ws1.iter_rows(min_row=3):  # Start from header row
        if any(all(row[df1.columns.get_loc(col)].value == unmatched_table1.iloc[i][col] 
            for col in common_columns) 
                for i in range(len(unmatched_table1))):
                    for cell in row:
                        cell.fill = red_fill
    
    # Save highlighted workbook
    wb1.save('highlighted_table1.xlsx')
    
    print(f"Unmatched in Table 1: {len(unmatched_table1)}")
    print(f"Unmatched in Table 2: {len(unmatched_table2)}")

# Get script directory
repo_path = os.path.dirname(os.path.abspath(__file__))

# Construct file paths
file1 = os.path.join(repo_path, '..', 'data', 'Collated transaction reports.xlsx')
file2 = os.path.join(repo_path, '..', 'data', 'Knowledge_Schools_Trust_-_Account_Transactions (23).xlsx')

# Run comparison and highlighting
compare_and_highlight_excel(file1, file2)