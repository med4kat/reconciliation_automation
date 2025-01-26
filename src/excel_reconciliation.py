

import pandas as pd
import os
import sys
import openpyxl

from openpyxl.styles import PatternFill

def compare_and_highlight_excel(file1, file2, sheet1=0, sheet2=0):
    # Read Excel files
    df1 = pd.read_excel(file1, sheet_name=sheet1, header=3)
    df2 = pd.read_excel(file2, sheet_name=sheet2, header=6)
    
    # # Find matching columns
    # common_columns = list(set(df1.columns) & set(df2.columns))

    common_columns = ['Date', 'Description', 'Debit']
 
    # Merge dataframes
    merged = pd.merge(df1, df2, 
                    on=common_columns, 
                    how='outer', 
                    indicator=True)
    
    
    
    # Identify unmatched transactions
    unmatched_table1 = merged[merged['_merge'] == 'left_only']
    unmatched_table2 = merged[merged['_merge'] == 'right_only']



    # Load workbooks
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)

    # Highlight unmatched in first file
    ws1 = wb1.worksheets[sheet1]
    ws2 = wb2.worksheets[sheet2]
    yelow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
    # Fantasy fill
    red_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
    # Fantasy Fill
    for row in ws1.iter_rows(min_row=3):  # Start from header row
        if any(all(row[df1.columns.get_loc(col)].value == unmatched_table1.iloc[i][col] for col in common_columns) for i in range(len(unmatched_table1))):
            for cell in row:
                cell.fill = yelow_fill

# Loop through each row in worksheet, starting at row 3
# for row in ws1.iter_rows(min_row=3):

# explain___________________________
    # Check if current row matches any unmatched transaction by comparing all columns
    # matches = False
    # for i in range(len(unmatched_table1)):
    #     all_cols_match = True
    #     for col in common_columns:
    #         if row[df1.columns.get_loc(col)].value != unmatched_table1.iloc[i][col]:
    #             all_cols_match = False
    #             break
    #     if all_cols_match:
    #         matches = True
    #         break

    # # If row matches an unmatched transaction, highlight all cells yellow
    # if matches:
    #     for cell in row:
    #         cell.fill = yellow_fill
# end of explain________________________
                        
    for row in ws2.iter_rows(min_row=6):
        if any(all(row[df2.columns.get_loc(col)].value == unmatched_table2.iloc[i][col] for col in common_columns) for i in range(len(unmatched_table2))):
            for cell in row:
                cell.fill = yelow_fill

    
    # Save highlighted workbook
    wb1.save('highlighted_table1.xlsx')
    wb2.save('highlighted_table2.xlsx')
    
    print(f"Unmatched in Table 1: {len(unmatched_table1)}")
    print(f"Unmatched in Table 2: {len(unmatched_table2)}")

# Get script directory
repo_path = os.path.dirname(os.path.abspath(__file__))

# Construct file paths
file1 = os.path.join(repo_path, '..', 'data', 'Collated transaction reports.xlsx')
file2 = os.path.join(repo_path, '..', 'data', 'Knowledge_Schools_Trust_-_Account_Transactions (23).xlsx')

# Run comparison and highlighting
compare_and_highlight_excel(file1, file2)